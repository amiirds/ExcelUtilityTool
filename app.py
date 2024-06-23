import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def read_and_format(input_filename, output_filename):
    try:
        with open(input_filename, 'r') as file:
            lines = file.read().splitlines()
        formatted_output = ', '.join([line for line in lines])

        if os.path.isdir(output_filename):
            print(f"Error: The output path {output_filename} is a directory. Please provide a file path.")
            return

        with open(output_filename, 'w') as file:
            file.write(formatted_output)
        print(f'Formatted list has been written to {output_filename}')
    except Exception as e:
        print(f"An error occurred: {e}")


def create_hyperlink_excel(input_file, output_file):
    try:
        if not os.path.exists(input_file):
            print(f"Error: The file {input_file} does not exist.")
            return
        wb = load_workbook(input_file)
        ws = wb.active
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
            name = row[0] if row[0] is not None else ''
            path = row[1] if row[1] is not None else ''
            link_cell = ws.cell(row=idx, column=1)
            link_cell.value = name
            if path and os.path.exists(path):
                link_cell.hyperlink = path
                link_cell.font = Font(color="0000FF", underline="single")
            elif path:
                print(f"Warning: The path {path} does not exist.")
        wb.save(output_file)
        print(f"Hyperlink Excel file created: {output_file}")
    except Exception as e:
        print(f"An error occurred: {e}")


def find_file_paths(directory, file_names):
    file_names_lower = [name.lower() for name in file_names]
    file_paths_dict = {name: [] for name in file_names}

    print(f"Scanning directory: {directory}")
    if not os.path.exists(directory):
        print(f"Error: The directory {directory} does not exist.")
        return file_paths_dict

    for root, _, files in os.walk(directory):
        print(f"Checking directory: {root}")
        for file_name in files:
            file_name_lower = file_name.lower()
            print(f"Found file: {file_name_lower}")
            for search_name in file_names_lower:
                if search_name in file_name_lower:
                    original_name = file_names[file_names_lower.index(search_name)]
                    file_path = os.path.join(root, file_name)
                    file_paths_dict[original_name].append(file_path)
                    print(f"Matched {original_name}: {file_path}")

    for name, paths in file_paths_dict.items():
        if not paths:
            print(f"File not found: {name}")

    return file_paths_dict


def save_paths_to_excel(file_paths_dict, output_directory):
    wb = Workbook()
    ws = wb.active
    row = 1
    for name, paths in file_paths_dict.items():
        sorted_paths = sorted(paths)
        ws.cell(row=row, column=1, value=name)
        for i, file_path in enumerate(sorted_paths, start=2):
            ws.cell(row=row, column=i, value=file_path)
        row += 1
    excel_file_path = os.path.join(output_directory, "sorted_file_paths.xlsx")
    try:
        wb.save(excel_file_path)
        print("Sorted file paths saved to:", excel_file_path)
    except Exception as e:
        print(f"Error saving Excel file: {e}")


def menu():
    while True:
        print("\nMenu:")
        print("1. Read and format a list of names")
        print("2. Create an Excel file with hyperlinks")
        print("3. Find file paths and save to Excel")
        print("4. Exit")
        choice = input("Enter your choice (1-4): ")

        if choice == '1':
            input_file = input("Enter the path to the input file: ")
            output_file = input("Enter the path to the output file (including file name): ")
            read_and_format(input_file, output_file)

        elif choice == '2':
            input_file = input("Enter the path to the input Excel file: ")
            output_file = input("Enter the path to the output Excel file (including file name): ")
            create_hyperlink_excel(input_file, output_file)

        elif choice == '3':
            directory = input("Enter the directory to search: ")
            file_names = input("Enter the file names to search (comma-separated): ").split(',')
            file_names = [name.strip() for name in file_names]
            file_paths_dict = find_file_paths(directory, file_names)
            if not any(file_paths_dict.values()):
                print("No files were found.")
            else:
                print("\nFound files:")
                for key, paths in file_paths_dict.items():
                    if paths:
                        for path in paths:
                            print(f"{key}: {path}")
                    else:
                        print(f"{key}: Not found")
                output_directory = input("Enter the directory to save the output Excel file: ")
                save_paths_to_excel(file_paths_dict, output_directory)

        elif choice == '4':
            print("Exiting the program.")
            break

        else:
            print("Invalid choice. Please try again.")


if __name__ == "__main__":
    menu()
